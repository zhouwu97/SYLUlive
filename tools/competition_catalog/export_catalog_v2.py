"""从工作簿或 CSV 离线导出 Catalog 2.2 JSON。"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

from _catalog_v2 import HASH_FIELDS, build_document, validate_document

GOVERNED_SHEETS = {"目录发布清单", "赛事基础导出", "AI预览视图"}


def read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def read_workbook(path: Path, sheet_name: str | None) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("读取 .xlsx 需要安装 openpyxl") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name:
        return read_worksheet(workbook[sheet_name], 1)
    if GOVERNED_SHEETS.issubset(workbook.sheetnames):
        rows, _ = read_governed_workbook(workbook)
        return rows
    return read_worksheet(workbook.active, 1)


def read_worksheet(worksheet: Any, header_row: int) -> list[dict[str, Any]]:
    rows = worksheet.iter_rows(values_only=True)
    for _ in range(header_row - 1):
        next(rows, None)
    try:
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    except StopIteration:
        return []
    if not all(headers) or len(headers) != len(set(headers)):
        raise ValueError("表头不能为空且不能重复")
    return [
        dict(zip(headers, values, strict=True))
        for values in rows
        if any(value is not None and str(value).strip() for value in values)
    ]


def read_parent_mapping(worksheet: Any, header_row: int) -> dict[str, str]:
    rows = worksheet.iter_rows(values_only=True)
    for _ in range(header_row - 1):
        next(rows, None)
    headers = [str(value).strip() if value is not None else "" for value in next(rows, ())]
    try:
        competition_id_index = headers.index("competition_id")
        parent_id_index = headers.index("parent_competition_id")
    except ValueError:
        return {}
    result: dict[str, str] = {}
    for values in rows:
        competition_id = str(values[competition_id_index] or "").strip()
        if competition_id:
            result[competition_id] = str(values[parent_id_index] or "").strip()
    return result


def read_governed_workbook(workbook: Any) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    manifest_rows = read_worksheet(workbook["目录发布清单"], 4)
    if len(manifest_rows) != 1:
        raise ValueError("目录发布清单必须且只能包含一条元数据记录")
    manifest = manifest_rows[0]
    if manifest.get("schema_version") != "sylulive-competition-catalog/2.2":
        raise ValueError("工作簿目录 Schema 版本不兼容")
    if manifest.get("current_validation_all_passed") is not True:
        raise ValueError("工作簿实时校验未全部通过")
    if manifest.get("current_content_drift_detected") is True:
        raise ValueError("工作簿检测到发布层内容漂移")

    base_rows = read_worksheet(workbook["赛事基础导出"], 4)
    preview_rows = read_worksheet(workbook["AI预览视图"], 4)
    preview_by_id = {str(row.get("competition_id") or "").strip(): row for row in preview_rows}
    source_parent_by_id: dict[str, str] = {}
    source_sheet = str(manifest.get("source_sheet") or "").strip()
    if source_sheet and source_sheet in workbook.sheetnames:
        source_parent_by_id = read_parent_mapping(workbook[source_sheet], 1)
    if len(base_rows) != len(preview_rows) or len(preview_by_id) != len(preview_rows):
        raise ValueError("赛事基础导出与 AI 预览视图数量或主键不一致")

    records: list[dict[str, Any]] = []
    for order, base in enumerate(base_rows, 1):
        competition_id = str(base.get("competition_id") or "").strip()
        preview = preview_by_id.get(competition_id)
        if preview is None:
            raise ValueError(f"{competition_id}: AI 预览记录缺失")
        if base.get("dataset_version") != preview.get("dataset_version"):
            raise ValueError(f"{competition_id}: 数据集版本跨视图不一致")
        records.append(
            {
                "competition_id": competition_id,
                "parent_competition_id": str(
                    base.get("parent_competition_id")
                    or preview.get("parent_competition_id")
                    or source_parent_by_id.get(competition_id)
                    or ""
                ).strip(),
                "record_hash": "",
                "catalog_order": order,
                "title": preview.get("display_title") or base.get("title"),
                "subtitle": "",
                "summary": base.get("summary"),
                "description": base.get("description"),
                "primary_category_slug": preview.get("primary_category_slug"),
                "tags": preview.get("app_tags_json"),
                "competition_level": preview.get("competition_level"),
                "school_recognition_status": preview.get("school_recognition_status"),
                "school_recognition_grade": preview.get("school_recognition_grade"),
                "competition_rating": preview.get("manual_rating"),
                "importance_score": 0,
                "organizer": base.get("organizer"),
                "host_unit": base.get("host_unit"),
                "target_audience": base.get("target_audience"),
                "eligible_entry_years": preview.get("eligible_grades_json"),
                "eligible_colleges": preview.get("eligible_colleges_json"),
                "eligible_majors": preview.get("eligible_majors_json"),
                "participation_type": base.get("participation_type"),
                "team_size_min": base.get("team_size_min"),
                "team_size_max": base.get("team_size_max"),
                "registration_start": base.get("registration_start"),
                "registration_end": base.get("registration_end"),
                "event_start": base.get("event_start"),
                "event_end": base.get("event_end"),
                "registration_time_text": base.get("registration_time_text"),
                "event_time_text": base.get("event_time_text"),
                "time_precision": base.get("time_precision"),
                "time_status": base.get("time_status"),
                "time_note": base.get("time_note"),
                "sort_month": base.get("sort_month"),
                "location": base.get("location"),
                "is_online": base.get("is_online"),
                "official_url": preview.get("official_url") or base.get("official_url"),
                "notice_url": preview.get("notice_url") or base.get("notice_url"),
                "source_channel": base.get("source_channel"),
                "source_note": base.get("source_note"),
                "status": base.get("status"),
                "manual_rating_reason_public": preview.get("manual_rating_reason_public"),
                "major_fit_summary_public": preview.get("major_fit_summary_public"),
                "evidence_summary_public": preview.get("evidence_summary_public"),
                "evidence_subgrade": preview.get("student_evidence_subgrade"),
                "risk_tags": preview.get("risk_tags_json"),
                "search_display_allowed": preview.get("search_display_allowed"),
                "candidate_pool_allowed": preview.get("candidate_pool_allowed"),
                "personalized_ranking_allowed": preview.get("personalized_ranking_allowed"),
                "strong_recommendation_eligible": preview.get("strong_recommendation_eligible"),
                "recommendation_permission_level": preview.get("recommendation_permission_level"),
                "ai_mode": preview.get("ai_mode"),
                "blocker_codes": preview.get("blocker_codes_json"),
            }
        )
    if any(set(record) != set(HASH_FIELDS) for record in records):
        raise ValueError("工作簿字段映射与 Go Catalog DTO 不一致")
    if int(manifest.get("item_count") or -1) != len(records):
        raise ValueError("目录发布清单 item_count 与导出记录数不一致")
    return records, manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, help="输入 .xlsx 或 .csv")
    parser.add_argument("output", type=Path, help="输出 JSON")
    parser.add_argument("--dataset-version")
    parser.add_argument("--publish-status", choices=("draft", "published"))
    parser.add_argument("--production-load-allowed", action="store_true")
    parser.add_argument("--sheet", help="工作表名称；默认使用活动工作表")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    manifest: dict[str, Any] | None = None
    if args.input.suffix.lower() == ".csv":
        rows = read_csv(args.input)
    elif args.input.suffix.lower() == ".xlsx":
        if args.sheet:
            rows = read_workbook(args.input, args.sheet)
        else:
            from openpyxl import load_workbook

            workbook = load_workbook(args.input, read_only=True, data_only=True)
            if GOVERNED_SHEETS.issubset(workbook.sheetnames):
                rows, manifest = read_governed_workbook(workbook)
            else:
                rows = read_workbook(args.input, None)
    else:
        raise ValueError("输入文件必须是 .xlsx 或 .csv")
    dataset_version = args.dataset_version or (
        str(manifest.get("dataset_version") or "").strip() if manifest else ""
    )
    publish_status = args.publish_status or (
        str(manifest.get("publish_status") or "").strip() if manifest else "draft"
    )
    if not dataset_version:
        raise ValueError("必须通过 --dataset-version 或工作簿清单提供数据集版本")
    production_load_allowed = args.production_load_allowed
    if manifest:
        if args.dataset_version and args.dataset_version != manifest.get("dataset_version"):
            raise ValueError("--dataset-version 与工作簿目录发布清单不一致")
        if args.publish_status and args.publish_status != manifest.get("publish_status"):
            raise ValueError("--publish-status 与工作簿目录发布清单不一致")
        production_load_allowed = bool(manifest.get("production_load_allowed"))
        if args.production_load_allowed and not production_load_allowed:
            raise ValueError("当前工作簿未开放 production_load_allowed")
    document = build_document(
        rows,
        dataset_version=dataset_version,
        publish_status=publish_status,
        production_load_allowed=production_load_allowed,
        source_filename=args.input.name,
    )
    errors = validate_document(document)
    if errors:
        raise ValueError("\n".join(errors))
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(document, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    print(
        json.dumps(
            {
                "status": "ok",
                "output": str(args.output),
                "item_count": document["item_count"],
                "package_hash": document["package_hash"],
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
