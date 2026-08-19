"""分层竞赛工作簿导出回归测试。"""

from __future__ import annotations

import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))

from export_catalog_v2 import read_governed_workbook


DATASET_VERSION = "2026.07.22-v8.2-activation"


def _sheet(workbook: Workbook, title: str, headers: list[str], values: list[object]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(["说明"])
    sheet.append(["说明"])
    sheet.append(["说明"])
    sheet.append(headers)
    sheet.append(values)


def _workbook(*, drift: bool = False, preview_id: str = "COMP-001") -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _sheet(
        workbook,
        "目录发布清单",
        [
            "schema_version",
            "dataset_version",
            "source_sheet",
            "publish_status",
            "production_load_allowed",
            "item_count",
            "current_validation_all_passed",
            "current_content_drift_detected",
        ],
        [
            "sylulive-competition-catalog/2.2",
            DATASET_VERSION,
            "导入清洗版",
            "draft",
            False,
            1,
            True,
            drift,
        ],
    )
    source = workbook.create_sheet("导入清洗版")
    source.append(["competition_id", "parent_competition_id"])
    source.append(["COMP-001", "COMP-PARENT"])
    _sheet(
        workbook,
        "赛事基础导出",
        [
            "competition_id",
            "dataset_version",
            "title",
            "summary",
            "description",
            "status",
            "time_precision",
            "time_status",
        ],
        [
            "COMP-001",
            DATASET_VERSION,
            "基础标题",
            "摘要",
            "描述",
            "published",
            "unknown",
            "pending",
        ],
    )
    _sheet(
        workbook,
        "AI预览视图",
        [
            "competition_id",
            "dataset_version",
            "display_title",
            "app_tags_json",
            "eligible_grades_json",
            "eligible_colleges_json",
            "eligible_majors_json",
            "risk_tags_json",
            "blocker_codes_json",
            "search_display_allowed",
            "candidate_pool_allowed",
            "personalized_ranking_allowed",
            "strong_recommendation_eligible",
            "recommendation_permission_level",
            "ai_mode",
        ],
        [
            preview_id,
            DATASET_VERSION,
            "预览标题",
            "[]",
            "[]",
            "[]",
            "[]",
            "[]",
            "[]",
            True,
            True,
            False,
            False,
            "low",
            "candidate_explanation",
        ],
    )
    return workbook


class GovernedWorkbookExportTest(unittest.TestCase):
    def test_merges_base_and_preview_views(self) -> None:
        records, manifest = read_governed_workbook(_workbook())

        self.assertEqual(manifest["publish_status"], "draft")
        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["competition_id"], "COMP-001")
        self.assertEqual(records[0]["title"], "预览标题")
        self.assertEqual(records[0]["summary"], "摘要")
        self.assertEqual(records[0]["parent_competition_id"], "COMP-PARENT")
        self.assertIs(records[0]["candidate_pool_allowed"], True)

    def test_rejects_content_drift(self) -> None:
        with self.assertRaisesRegex(ValueError, "内容漂移"):
            read_governed_workbook(_workbook(drift=True))

    def test_rejects_missing_cross_view_record(self) -> None:
        with self.assertRaisesRegex(ValueError, "AI 预览记录缺失"):
            read_governed_workbook(_workbook(preview_id="COMP-OTHER"))

    def test_cli_cannot_override_draft_manifest_as_published(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            temp_path = Path(directory)
            workbook_path = temp_path / "catalog.xlsx"
            output_path = temp_path / "catalog.json"
            _workbook().save(workbook_path)

            result = subprocess.run(
                [
                    sys.executable,
                    str(Path(__file__).with_name("export_catalog_v2.py")),
                    str(workbook_path),
                    str(output_path),
                    "--publish-status",
                    "published",
                ],
                check=False,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
            )

            self.assertNotEqual(result.returncode, 0)
            # Windows 子进程的控制台编码可能替换中文，仅断言稳定的参数标识。
            self.assertIn("--publish-status", result.stderr)
            self.assertFalse(output_path.exists())


if __name__ == "__main__":
    unittest.main()
